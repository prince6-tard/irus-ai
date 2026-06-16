"""FastAPI application for the IRUS AI Outreach Pipeline.

Bridges the Next.js frontend with the Python orchestration engine.
"""

from typing import Any

from fastapi import BackgroundTasks, FastAPI
from fastapi import UploadFile, File
from fastapi.middleware.cors import CORSMiddleware

import csv
import io
import re
import openpyxl

from config import DRY_RUN, DELAY_MIN_SECONDS, DELAY_MAX_SECONDS
from db import get_all_leads, get_lead_count, init_tables, get_leads_by_ids, init_dental_tables, get_dental_leads, get_dental_email_log, update_dental_lead_status, log_dental_email, get_dental_leads_by_ids, save_dental_lead, init_medical_tourism_tables
from claude_drafter import draft_email
from email_sender import send_email
from dental_scraper import run_dental_scrape
from product_loader import format_selected_catalogue
import random
import time
from logger import write_log
from email_drafts import save_draft, mark_sent, mark_failed
from email_drafts import get_all_drafts, get_sendable_drafts, update_draft
from logger import get_logs
from main import run_campaign, send_all, send_single
from product_loader import get_catalogue_json


# In-memory status tracker for the active campaign
_campaign_status: dict[str, Any] = {
    "running": False,
    "result": None,
}


def _run_campaign_task(payload: dict) -> None:
    """Wrapper that runs the campaign and updates in-memory status."""
    global _campaign_status
    _campaign_status["running"] = True
    _campaign_status["result"] = None
    try:
        result = run_campaign(payload)
        _campaign_status["result"] = result
    except Exception as exc:
        _campaign_status["result"] = {"status": "error", "message": str(exc)}
    finally:
        _campaign_status["running"] = False


# ── FastAPI App ────────────────────────────────────────────────────────────────

app = FastAPI(title="IRUS AI Outreach API", version="4.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _startup() -> None:
    try:
        init_tables()
    except Exception as exc:
        print(f"[startup] DB init warning: {exc}")
    try:
        init_dental_tables()
    except Exception as exc:
        print(f"[startup] Dental DB init warning: {exc}")
    try:
        init_medical_tourism_tables()
    except Exception as exc:
        print(f"[startup] MedTour DB init warning: {exc}")


# ── Products ───────────────────────────────────────────────────────────────────

@app.get("/products")
def get_products() -> dict:
    """Return the full product catalogue as JSON."""
    try:
        products = get_catalogue_json()
        return {"status": "success", "count": len(products), "products": products}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


# ── Campaign Launch ────────────────────────────────────────────────────────────

@app.post("/launch")
def launch_campaign(payload: dict, background_tasks: BackgroundTasks) -> dict:
    """Start the pipeline: discovery → enrichment → AI drafting.

    Expected payload:
        {
            "domain": "Defence" | "Medical",
            "locations": ["India", "UAE"],
            "selected_products": ["Product A", "Product B"],
            "dry_run": true
        }
    """
    if _campaign_status["running"]:
        return {
            "status": "busy",
            "message": "A campaign is already running. Please wait.",
        }

    required = {"domain", "locations", "selected_products"}
    missing = required - set(payload.keys())
    if missing:
        return {
            "status": "error",
            "message": f"Missing required fields: {', '.join(sorted(missing))}",
        }

    background_tasks.add_task(_run_campaign_task, payload)
    return {
        "status": "started",
        "message": "Campaign started in the background.",
        "payload": {k: payload[k] for k in required if k in payload},
    }


# ── Drafts ─────────────────────────────────────────────────────────────────────

@app.get("/drafts")
def get_drafts(status: str | None = None) -> dict:
    """Return all AI-drafted emails. Optional status filter: drafted, edited, sent, failed."""
    try:
        drafts = get_all_drafts(status=status)
        return {"status": "success", "count": len(drafts), "drafts": drafts}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.get("/drafts/pending")
def get_pending_drafts() -> dict:
    """Return all drafts that are ready to send (drafted or edited)."""
    try:
        drafts = get_sendable_drafts()
        return {"status": "success", "count": len(drafts), "drafts": drafts}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.put("/drafts/{email}")
def edit_draft(email: str, payload: dict) -> dict:
    """Update the subject and/or body of a drafted email. Sets status to 'edited'."""
    try:
        subject = payload.get("subject")
        body = payload.get("body")
        if subject is None and body is None:
            return {"status": "error", "message": "Nothing to update. Provide subject or body."}
        updated = update_draft(email, subject=subject, body=body)
        if updated:
            return {"status": "success", "message": f"Draft for {email} updated."}
        return {"status": "error", "message": f"Draft for {email} not found."}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.post("/drafts/{email}/send")
def send_one(email: str) -> dict:
    """Send a single reviewed email."""
    try:
        return send_single(email)
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.post("/drafts/send-all")
def send_all_drafts(background_tasks: BackgroundTasks) -> dict:
    """Send all pending (drafted/edited) emails in the background."""
    background_tasks.add_task(send_all)
    return {
        "status": "started",
        "message": "Sending all pending emails in the background.",
    }


# ── Upload ────────────────────────────────────────────────────────────────────

HEADER_SYNONYMS: dict[str, list[str]] = {
    "name":               ["name", "full name", "contact name", "person"],
    "email":              ["email", "e-mail", "email address"],
    "organization":       ["organization", "org", "company", "organisation", "firm"],
    "role":               ["role", "title", "job title", "designation"],
    "city":               ["city", "town"],
    "country":            ["country"],
    "phone":              ["phone", "mobile", "contact number", "phone number"],
    "category":           ["category", "domain", "industry"],
    "organization_domain":["domain", "website", "url"],
    "linkedin_url":       ["linkedin", "linkedin url"],
    "apollo_id":          ["apollo_id", "apollo id", "id"],
}

# Reverse: synonym -> standard key
_SYNONYM_MAP: dict[str, str] = {
    synonym: key
    for key, synonyms in HEADER_SYNONYMS.items()
    for synonym in synonyms
}


def _normalize_header(h: str) -> str:
    """Lowercase, strip spaces/underscores to match synonyms."""
    return h.strip().lower().replace(" ", "_").replace("_", "_")


def _map_header(header: str) -> str:
    """Map a raw header to a standard lead key, or None to drop."""
    normalized = _normalize_header(header)
    # Try exact synonym match first
    if normalized in _SYNONYM_MAP:
        return _SYNONYM_MAP[normalized]
    # Fall back: check if the raw header itself (lower/stripped) is a known key
    base = header.strip().lower()
    if base in _SYNONYM_MAP:
        return _SYNONYM_MAP[base]
    return None  # unknown header – skip


# Stricter email regex: prevents trailing concatenated words (e.g. .comrequest)
EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,6}$")


def _parse_csv(buffer: io.BytesIO) -> tuple[list[dict], int]:
    """Parse a CSV file. Returns (rows, skipped_count). Decodes with utf-8/latin-1 fallback."""
    raw = buffer.read()
    for enc in ("utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    header_map = {orig: _map_header(orig) for orig in reader.fieldnames or []}

    rows, skipped = [], 0
    for raw_row in reader:
        row = {}
        notes = []
        for orig_val, key in ((raw_row.get(k, "") or "", header_map.get(k))
                               for k in header_map):
            if key is None:
                if orig_val.strip():
                    notes.append(orig_val.strip())
            else:
                row[key] = orig_val.strip()

        if notes:
            row["notes"] = "; ".join(notes)

        email = row.get("email", "").strip()
        if not email or not EMAIL_RE.match(email):
            skipped += 1
            continue

        row["email"] = email
        row["source"] = "upload"
        rows.append(row)

    return rows, skipped


def _parse_xlsx(buffer: io.BytesIO) -> tuple[list[dict], int]:
    """Parse an XLSX file. Returns (rows, skipped_count)."""
    wb = openpyxl.load_workbook(buffer, data_only=True)
    ws = wb.active

    headers = [cell.value or "" for cell in next(ws.iter_rows(max_row=1))]
    header_map = {h: _map_header(h) for h in headers}

    rows, skipped = [], 0
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        notes = []
        for cell_val, key in ((v or "", header_map.get(h))
                               for v, h in zip(raw_row, headers)):
            if key is None:
                if str(cell_val).strip():
                    notes.append(str(cell_val).strip())
            else:
                row[key] = str(cell_val).strip()

        if notes:
            row["notes"] = "; ".join(notes)

        email = row.get("email", "").strip()
        if not email or not EMAIL_RE.match(email):
            skipped += 1
            continue

        row["email"] = email
        row["source"] = "upload"
        rows.append(row)

    return rows, skipped


@app.post("/upload")
async def upload_leads(file: UploadFile = File(...)) -> dict:
    """Accept a CSV or XLSX file, parse leads, validate emails, and insert into DB.

    Returns:
        {
            "status": "success" | "error",
            "inserted": int,
            "skipped": int,
            "leads": [lead_dict, ...]
        }
    """
    filename = file.filename or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext not in ("csv", "xlsx", "xls"):
        return {
            "status": "error",
            "message": "Unsupported file type. Please upload a .csv, .xlsx, or .xls file.",
            "inserted": 0,
            "skipped": 0,
            "leads": [],
        }

    try:
        buffer = io.BytesIO(await file.read())
    except Exception as exc:
        return {
            "status": "error",
            "message": f"Failed to read file: {exc}",
            "inserted": 0,
            "skipped": 0,
            "leads": [],
        }

    if ext == "csv":
        rows, skipped = _parse_csv(buffer)
    else:
        rows, skipped = _parse_xlsx(buffer)

    if not rows:
        return {
            "status": "success",
            "inserted": 0,
            "skipped": skipped,
            "leads": [],
        }

    try:
        from db import insert_leads
        inserted = insert_leads(rows)
    except Exception as exc:
        return {
            "status": "error",
            "message": f"Database insertion failed: {exc}",
            "inserted": 0,
            "skipped": skipped + len(rows),
            "leads": [],
        }

    return {
        "status": "success",
        "inserted": inserted,
        "skipped": skipped,
        "leads": rows,
    }


# ── Leads ──────────────────────────────────────────────────────────────────────

@app.get("/leads")
def get_leads(category: str | None = None, search: str | None = None, limit: int = 500, offset: int = 0) -> dict:
    """Return all scraped leads with optional filters."""
    try:
        leads = get_all_leads(category=category, search=search, limit=limit, offset=offset)
        total = get_lead_count(category=category, search=search)
        return {"status": "success", "count": len(leads), "total": total, "leads": leads}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.post("/leads/send-selected")
def send_selected_leads(payload: dict) -> dict:
    """Draft and send emails to selected leads for selected products.

    Body: { "lead_ids": [int], "selected_products": [str] }
    """
    try:
        lead_ids = payload.get("lead_ids", [])
        selected_products = payload.get("selected_products", [])

        if not lead_ids:
            return {"status": "error", "message": "lead_ids is required and must be non-empty"}
        if not selected_products:
            return {"status": "error", "message": "selected_products is required and must be non-empty"}

        leads = get_leads_by_ids(lead_ids)
        if not leads:
            return {"status": "error", "message": "No leads found for the provided IDs"}

        selected_products_text = format_selected_catalogue(selected_products)

        sent = 0
        failed = 0
        skipped = 0
        dry_run = payload.get("dry_run", DRY_RUN)

        for i, lead in enumerate(leads):
            email = lead.get("email", "").strip()
            if not email:
                write_log(lead, "skipped_no_email", note="No email address on lead")
                skipped += 1
                continue

            name = lead.get("name", "") or ""
            subject, body = draft_email(lead, selected_products_text)

            if dry_run:
                save_draft(lead, subject, body, status="drafted")
                write_log(lead, "dry_run", subject=subject)
                sent += 1
            else:
                try:
                    ok = send_email(email, name, subject, body)
                    if ok:
                        mark_sent(email)
                        write_log(lead, "sent", subject=subject)
                        sent += 1
                    else:
                        mark_failed(email)
                        write_log(lead, "error_send", subject=subject, note="send_email returned False")
                        failed += 1
                except Exception as exc:
                    mark_failed(email)
                    write_log(lead, "error_send", subject=subject, note=str(exc))
                    failed += 1

            # Delay between sends, but not after the last one
            if i < len(leads) - 1:
                time.sleep(random.randint(DELAY_MIN_SECONDS, DELAY_MAX_SECONDS))

        return {
            "status": "success",
            "sent": sent,
            "failed": failed,
            "skipped": skipped,
            "dry_run": dry_run,
        }
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


# ── Logs & Status ──────────────────────────────────────────────────────────────

@app.get("/logs")
def get_logs_endpoint(limit: int = 500) -> dict:
    try:
        rows = get_logs(limit=limit)
        return {"status": "success", "count": len(rows), "logs": rows}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.get("/status")
def get_status() -> dict:
    return {
        "status": "success",
        "running": _campaign_status["running"],
        "last_result": _campaign_status["result"],
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  Dental Connect Endpoints
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/dental/scrape")
async def dental_scrape(payload: dict) -> dict:
    """Trigger medical tourism facilitator scraping for a list of queries."""
    queries = payload.get("queries", [])
    if not queries:
        return {"status": "error", "message": "No queries provided"}
    result = run_dental_scrape(queries)
    return {"status": "ok", **result}


@app.get("/dental/leads")
def dental_leads(status: str | None = None) -> dict:
    """Return dental leads, optionally filtered by status."""
    try:
        leads = get_dental_leads(status=status)
        return {"status": "success", "count": len(leads), "leads": leads}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.get("/dental/log")
def dental_log() -> dict:
    """Return dental email send log."""
    try:
        logs = get_dental_email_log()
        return {"status": "success", "count": len(logs), "logs": logs}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.post("/dental/send")
def dental_send(payload: dict) -> dict:
    """Send Dr. Swati Singhal collaboration emails to selected dental leads.

    Body: { "lead_ids": [int] }
    """
    lead_ids = payload.get("lead_ids", [])
    if not lead_ids:
        return {"status": "error", "message": "lead_ids is required"}

    leads = get_dental_leads_by_ids(lead_ids)
    if not leads:
        return {"status": "error", "message": "No leads found for provided IDs"}

    sent = 0
    failed = 0
    results = []

    for lead in leads:
        lead_id = lead.get("id")
        email = (lead.get("email") or "").strip()
        name = lead.get("name", "")
        organization = lead.get("organization", "")
        city = lead.get("city", "")
        country = lead.get("country", "")

        if not email:
            results.append({"lead_id": lead_id, "email": "", "status": "failed"})
            failed += 1
            continue

        subject = "Partnership Opportunity – Referral Collaboration with Spreading Smiles Dental Center"
        body = f"""Dear {organization or name or "Recipient Name"},

I hope you are doing well.

My name is Dr. Swati Singhal, Founder of Spreading Smiles Dental Center, India.

We frequently work with international patients who require extensive dental treatments such as implants, full-mouth rehabilitation, smile makeovers, crowns, bridges, and complex restorative procedures but are unable to proceed due to treatment costs in their home countries.

We would like to explore a referral partnership with your clinic for patients who may be seeking more affordable treatment alternatives without compromising quality or safety.

Our center offers:

* Comprehensive dental treatment by experienced specialists
* Significant cost savings compared to North American treatment costs
* Pre-treatment virtual consultations
* Digital treatment planning and cost estimates
* Dedicated patient coordinators
* Assistance with travel and accommodation
* Secure online access to treatment records
* AI-powered multilingual patient communication
* Post-treatment follow-up through our mobile care platform
* Ongoing support after patients return home

Our goal is not simply to provide treatment but to create a transparent, predictable, and well-coordinated patient experience from the first consultation to final follow-up.

We would welcome the opportunity to discuss how we can support your patients and establish a mutually beneficial referral relationship.

Thank you for your consideration, and we look forward to hearing from you.

Warm Regards,

Dr. Swati Singhal
Founder – Spreading Smiles Dental Center

📞 +91 9560035488
🌐 http://www.ihp.ind.in/spreading-smiles"""

        try:
            ok = send_email(email, name or organization, subject, body)
            if ok:
                update_dental_lead_status(lead_id, "sent")
                log_dental_email(lead_id, email, name, organization, city, country, "sent")
                sent += 1
                results.append({"lead_id": lead_id, "email": email, "status": "sent"})
            else:
                log_dental_email(lead_id, email, name, organization, city, country, "failed")
                failed += 1
                results.append({"lead_id": lead_id, "email": email, "status": "failed"})
        except Exception as exc:
            log_dental_email(lead_id, email, name, organization, city, country, "failed")
            failed += 1
            results.append({"lead_id": lead_id, "email": email, "status": "failed", "error": str(exc)})

    return {"status": "success", "sent": sent, "failed": failed, "results": results}


@app.post("/dental/upload")
async def dental_upload(file: UploadFile = File(...)) -> dict:
    """Upload CSV/XLSX file and save leads into dental_leads table.

    Parses the file using the same header-synonym mapping as /upload,
    then inserts each valid row into dental_leads with status='new'.
    """
    filename = file.filename or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext not in ("csv", "xlsx", "xls"):
        return {
            "status": "error",
            "message": "Unsupported file type. Please upload a .csv, .xlsx, or .xls file.",
            "inserted": 0,
            "skipped": 0,
            "leads": [],
        }

    try:
        buffer = io.BytesIO(await file.read())
    except Exception as exc:
        return {
            "status": "error",
            "message": f"Failed to read file: {exc}",
            "inserted": 0,
            "skipped": 0,
            "leads": [],
        }

    if ext == "csv":
        rows, skipped = _parse_csv(buffer)
    else:
        rows, skipped = _parse_xlsx(buffer)

    if not rows:
        return {
            "status": "success",
            "inserted": 0,
            "skipped": skipped,
            "leads": [],
        }

    inserted = 0
    saved_leads = []
    for row in rows:
        dental_lead = {
            "name": row.get("name", ""),
            "email": row.get("email", ""),
            "phone": row.get("phone", ""),
            "organization": row.get("organization", ""),
            "city": row.get("city", ""),
            "country": row.get("country", ""),
            "organization_domain": row.get("organization_domain", ""),
            "status": "new",
        }
        if save_dental_lead(dental_lead):
            inserted += 1
            saved_leads.append(dental_lead)

    return {
        "status": "success",
        "inserted": inserted,
        "skipped": skipped + (len(rows) - inserted),
        "leads": saved_leads,
    }


@app.patch("/dental/leads/status")
def dental_update_status(payload: dict) -> dict:
    """Bulk-update dental lead status.

    Body: { "lead_ids": [int], "status": "skipped" | "new" | "sent" }
    """
    lead_ids = payload.get("lead_ids", [])
    status = payload.get("status", "")
    if not lead_ids or not status:
        return {"status": "error", "message": "lead_ids and status are required"}

    updated = 0
    for lead_id in lead_ids:
        if update_dental_lead_status(lead_id, status):
            updated += 1

    return {"status": "success", "updated": updated}
